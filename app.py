"""
app.py — BioField Dashboard  (v4 — Cloud Edition)
author: Dr. Hakim Mitiche  |  Flask UI by Claude

Storage:
  - Images  → Cloudinary (object storage, permanent URLs)
  - Data    → PostgreSQL  (all observations, metadata, processing state)

No local files required. Works on Render, Railway, Fly.io, etc.
"""


from __future__ import annotations

import csv
import io
import json
import logging
import os
import queue
import subprocess
import sys
import threading
import zipfile
from collections import Counter, defaultdict
from functools import wraps
from pathlib import Path

import cloudinary
import cloudinary.uploader
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv
from flask import (Flask, Response, abort, jsonify, redirect, render_template,
                   request, send_file, stream_with_context, url_for)
from flask_login import (LoginManager, UserMixin, current_user, login_required,
                          login_user, logout_user)
from werkzeug.security import check_password_hash, generate_password_hash

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024   # 50 MB per upload
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "biofield-dev-secret-change-in-prod")

# ── Config ─────────────────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).parent

CATEGORIES = {
    "insect": {"label": "Insects", "color": "#f59e0b", "icon": "🪲"},
    "flora":  {"label": "Flora",   "color": "#4ade80", "icon": "🌿"},
    "fungus": {"label": "Fungi",   "color": "#c084fc", "icon": "🍄"},
}

ALLOWED_EXT = {".jpg", ".jpeg", ".png", ".webp"}

# ── Cloudinary ─────────────────────────────────────────────────────────────────

cloudinary.config(
    cloud_name = os.environ["CLOUDINARY_CLOUD_NAME"],
    api_key    = os.environ["CLOUDINARY_API_KEY"],
    api_secret = os.environ["CLOUDINARY_API_SECRET"],
    secure     = True,
)

# ── PostgreSQL ─────────────────────────────────────────────────────────────────

def get_db():
    """Return a new psycopg2 connection. Use as a context manager."""
    return psycopg2.connect(os.environ["DATABASE_URL"], sslmode="require")


def init_db():
    """Create tables and indexes if they don't exist. Each statement runs separately."""
    statements = [
        """
        CREATE TABLE IF NOT EXISTS observations (
            id                SERIAL PRIMARY KEY,
            category          VARCHAR(20)  NOT NULL,
            picture_name      VARCHAR(255),
            cloudinary_url    TEXT,
            cloudinary_public_id VARCHAR(255),
            common_name       VARCHAR(255),
            scientific_name   VARCHAR(255),
            species_name      VARCHAR(255),
            date              VARCHAR(50),
            gps_string        TEXT,
            latitude_dd       DOUBLE PRECISION,
            longitude_dd      DOUBLE PRECISION,
            altitude_m        DOUBLE PRECISION,
            processing_status VARCHAR(50)  DEFAULT 'PENDING',
            created_at        TIMESTAMP    DEFAULT NOW()
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS users (
            id            SERIAL PRIMARY KEY,
            username      VARCHAR(100) UNIQUE NOT NULL,
            email         VARCHAR(255) UNIQUE NOT NULL,
            password_hash VARCHAR(255) NOT NULL,
            role          VARCHAR(20)  DEFAULT 'user',
            created_at    TIMESTAMP    DEFAULT NOW()
        )
        """,
        "CREATE INDEX IF NOT EXISTS idx_obs_category ON observations(category)",
        "CREATE INDEX IF NOT EXISTS idx_obs_status   ON observations(processing_status)",
        # Safe to run on existing deployments — adds column only if missing
        "ALTER TABLE observations ADD COLUMN IF NOT EXISTS ai_confidence VARCHAR(20)",
    ]
    conn = get_db()
    try:
        with conn.cursor() as cur:
            for sql in statements:
                cur.execute(sql)
        conn.commit()
        app.logger.info("Database initialised.")
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def seed_admin():
    """Create the admin user from env vars if it doesn't exist yet."""
    username = os.environ.get("ADMIN_USERNAME", "admin")
    email    = os.environ.get("ADMIN_EMAIL",    "admin@biofield.local")
    password = os.environ.get("ADMIN_PASSWORD", "")

    if not password:
        app.logger.warning("ADMIN_PASSWORD not set — admin account not seeded.")
        return

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM users WHERE username = %s", (username,))
            if cur.fetchone():
                return
            cur.execute("""
                INSERT INTO users (username, email, password_hash, role)
                VALUES (%s, %s, %s, 'admin')
            """, (username, email, generate_password_hash(password)))
        conn.commit()
    app.logger.info(f"Admin user '{username}' created.")


# ── Auth ───────────────────────────────────────────────────────────────────────

login_manager = LoginManager(app)
login_manager.login_view = "login"
login_manager.login_message = ""


class User(UserMixin):
    def __init__(self, id, username, email, role):
        self.id = id
        self.username = username
        self.email = email
        self.role = role

    @property
    def is_admin(self):
        return self.role == "admin"


@login_manager.user_loader
def load_user(user_id):
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id, username, email, role FROM users WHERE id = %s",
                    (user_id,)
                )
                row = cur.fetchone()
        if not row:
            return None
        return User(*row)
    except Exception:
        return None


@login_manager.unauthorized_handler
def unauthorized():
    """Return JSON 401 for API calls, redirect to login for page requests."""
    if request.path.startswith("/api/"):
        return jsonify({"error": "Authentication required"}), 401
    return redirect(url_for("login", next=request.url))


def admin_required(f):
    """Decorator: require admin role. Returns 403 JSON for API, redirects for pages."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            if request.path.startswith("/api/"):
                return jsonify({"error": "Authentication required"}), 401
            return redirect(url_for("login", next=request.url))
        if not current_user.is_admin:
            if request.path.startswith("/api/"):
                return jsonify({"error": "Admin access required"}), 403
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return decorated


# ── Data Helpers ───────────────────────────────────────────────────────────────

def get_all_observations(category: str = "all") -> list[dict]:
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            if category == "all":
                cur.execute("SELECT * FROM observations ORDER BY created_at DESC")
            else:
                cur.execute(
                    "SELECT * FROM observations WHERE category = %s ORDER BY created_at DESC",
                    (category,)
                )
            rows = cur.fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d["_category"] = d["category"]
        result.append(d)
    return result


def get_stats() -> dict:
    stats = {"total": 0, "with_gps": 0, "categories": {}, "db_processed": 0}

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT category,
                       COUNT(*)                                          AS total,
                       COUNT(*) FILTER (WHERE gps_string IS NOT NULL
                                          AND gps_string != 'No GPS Data'
                                          AND gps_string != '')          AS with_gps,
                       COUNT(*) FILTER (WHERE cloudinary_url IS NOT NULL) AS images_in_cloud
                FROM observations
                GROUP BY category
            """)
            for cat, total, gps, imgs in cur.fetchall():
                stats["categories"][cat] = {
                    "total":          total,
                    "with_gps":       gps,
                    "images_on_disk": imgs,
                    "label": CATEGORIES.get(cat, {}).get("label", cat),
                    "color": CATEGORIES.get(cat, {}).get("color", "#888"),
                    "icon":  CATEGORIES.get(cat, {}).get("icon",  "?"),
                }
                stats["total"]    += total
                stats["with_gps"] += gps

            for cat, cfg in CATEGORIES.items():
                if cat not in stats["categories"]:
                    stats["categories"][cat] = {
                        "total": 0, "with_gps": 0, "images_on_disk": 0,
                        "label": cfg["label"], "color": cfg["color"], "icon": cfg["icon"],
                    }

            cur.execute("SELECT COUNT(*) FROM observations WHERE processing_status = 'SUCCESS'")
            stats["db_processed"] = cur.fetchone()[0]

    stats["map_exists"]     = (BASE_DIR / "bio_observations_map.html").exists()
    stats["geojson_exists"] = (BASE_DIR / "bio_observations.geojson").exists()
    return stats


def get_recent_observations(limit: int = 10) -> list[dict]:
    with get_db() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                "SELECT * FROM observations ORDER BY created_at DESC LIMIT %s",
                (limit,)
            )
            rows = cur.fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d["_category"] = d["category"]
        result.append(d)
    return result


def get_analytics_data() -> dict:
    rows = get_all_observations()

    month_counts: dict[str, int] = defaultdict(int)
    for r in rows:
        d = r.get("date", "") or ""
        if d and d != "N/A" and len(d) >= 7:
            month_counts[d[:7]] += 1
    months_sorted = sorted(month_counts.items())

    species_by_cat: dict[str, set] = defaultdict(set)
    for r in rows:
        name = (r.get("common_name") or r.get("species_name") or "").strip()
        if name and name.lower() not in ("unknown", ""):
            species_by_cat[r["_category"]].add(name)

    status_counts: Counter = Counter()
    for r in rows:
        st = r.get("processing_status") or "UNKNOWN"
        status_counts[st] += 1

    alt_buckets = {"0–500 m": 0, "500–1000 m": 0, "1000–1500 m": 0, "1500+ m": 0}
    for r in rows:
        try:
            a = float(r.get("altitude_m") or 0)
            if a < 500:       alt_buckets["0–500 m"]    += 1
            elif a < 1000:    alt_buckets["500–1000 m"] += 1
            elif a < 1500:    alt_buckets["1000–1500 m"]+= 1
            else:             alt_buckets["1500+ m"]    += 1
        except (ValueError, TypeError):
            pass

    all_species: Counter = Counter()
    for r in rows:
        name = (r.get("common_name") or r.get("species_name") or "").strip()
        if name and name.lower() not in ("unknown", ""):
            all_species[name] += 1

    with_gps    = sum(1 for r in rows
                      if (r.get("gps_string") or "").strip() not in ("", "No GPS Data"))
    without_gps = len(rows) - with_gps

    return {
        "total": len(rows),
        "timeline": {
            "labels": [m[0] for m in months_sorted],
            "values": [m[1] for m in months_sorted],
        },
        "species_richness": {cat: len(sp) for cat, sp in species_by_cat.items()},
        "status":           dict(status_counts),
        "altitude":         alt_buckets,
        "top_species":      all_species.most_common(10),
        "gps_coverage":     {"With GPS": with_gps, "No GPS": without_gps},
    }


# ── SSE Streaming ──────────────────────────────────────────────────────────────

_log_queues: dict[str, queue.Queue] = {}


def _stream_subprocess(cmd: list[str], stream_id: str):
    q = _log_queues[stream_id]
    try:
        env = {**os.environ}
        proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                text=True, bufsize=1, cwd=str(BASE_DIR), env=env)
        for line in proc.stdout:
            q.put({"type": "log", "text": line.rstrip()})
        proc.wait()
        q.put({"type": "done", "rc": proc.returncode,
               "text": f"[Exit code {proc.returncode}]"})
    except Exception as e:
        q.put({"type": "error", "text": str(e)})
    finally:
        q.put(None)


def sse_generator(stream_id: str):
    q = _log_queues.get(stream_id)
    if not q:
        yield "data: {}\n\n"
        return
    while True:
        item = q.get()
        if item is None:
            yield "data: " + json.dumps({"type": "end"}) + "\n\n"
            _log_queues.pop(stream_id, None)
            break
        yield "data: " + json.dumps(item) + "\n\n"


# ── Global context ─────────────────────────────────────────────────────────────

@app.context_processor
def inject_globals():
    return {"categories": CATEGORIES}


# ── Setup / Migration (no login required, token-protected) ────────────────────

@app.route("/api/setup")
def api_setup():
    """
    One-time migration endpoint. Call with ?token=<SETUP_TOKEN> to create
    missing tables and seed the admin. Safe to call multiple times.
    """
    token    = request.args.get("token", "")
    expected = os.environ.get("SETUP_TOKEN", "")
    if not expected:
        return jsonify({"error": "SETUP_TOKEN env var not set on the server."}), 403
    if token != expected:
        return jsonify({"error": "Invalid token."}), 403

    results = []
    try:
        init_db()
        results.append("Tables created / verified.")
    except Exception as e:
        return jsonify({"error": f"init_db failed: {e}"}), 500

    try:
        seed_admin()
        results.append("Admin user seeded (or already exists).")
    except Exception as e:
        results.append(f"seed_admin warning: {e}")

    return jsonify({"status": "ok", "steps": results})


# ── Auth Routes ────────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("index"))

    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id, username, email, role, password_hash FROM users WHERE username = %s",
                    (username,)
                )
                row = cur.fetchone()
        if row and check_password_hash(row[4], password):
            user = User(row[0], row[1], row[2], row[3])
            login_user(user, remember=True)
            next_page = request.args.get("next", "")
            if not next_page or not next_page.startswith("/"):
                next_page = url_for("index")
            return redirect(next_page)
        error = "Invalid username or password."

    return render_template("login.html", error=error)


@app.route("/register", methods=["GET", "POST"])
def register():
    if current_user.is_authenticated:
        return redirect(url_for("index"))

    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        email    = request.form.get("email", "").strip()
        password = request.form.get("password", "")
        confirm  = request.form.get("confirm", "")

        if not username or not email or not password:
            error = "All fields are required."
        elif len(password) < 6:
            error = "Password must be at least 6 characters."
        elif password != confirm:
            error = "Passwords do not match."
        else:
            try:
                with get_db() as conn:
                    with conn.cursor() as cur:
                        cur.execute("""
                            INSERT INTO users (username, email, password_hash, role)
                            VALUES (%s, %s, %s, 'user')
                        """, (username, email, generate_password_hash(password)))
                    conn.commit()
                return redirect(url_for("login"))
            except psycopg2.IntegrityError:
                error = "Username or email already taken."
            except Exception as e:
                error = f"Registration failed: {e}"

    return render_template("register.html", error=error)


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


# ── Page Routes ────────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def index():
    return render_template("index.html", stats=get_stats(), recent=get_recent_observations(8))


@app.route("/observations")
@login_required
def observations():
    cat_filter = request.args.get("category", "all")
    all_obs    = get_all_observations(cat_filter)
    return render_template("observations.html", observations=all_obs,
                           active=cat_filter, total=len(all_obs))


@app.route("/map")
def map_view():
    """Public — accessible without login."""
    map_file = BASE_DIR / "bio_observations_map.html"
    geo_file = BASE_DIR / "bio_observations.geojson"
    return render_template("map_view.html",
                           map_exists=map_file.exists(),
                           geojson_exists=geo_file.exists())


@app.route("/map-content")
def map_content():
    """Public — serves the Folium HTML file."""
    map_file = BASE_DIR / "bio_observations_map.html"
    if not map_file.exists():
        return "Map not generated yet.", 404
    return map_file.read_text(encoding="utf-8"), 200, {"Content-Type": "text/html"}


@app.route("/analytics")
@login_required
def analytics():
    return render_template("analytics.html", data=get_analytics_data())


@app.route("/upload")
@login_required
def upload_page():
    stats = get_stats()
    cats_with_counts = {}
    for cat, cfg in CATEGORIES.items():
        cats_with_counts[cat] = dict(cfg)
        cats_with_counts[cat]["images_on_disk"] = stats["categories"].get(
            cat, {}).get("images_on_disk", 0)
    return render_template("upload.html", categories=cats_with_counts)


@app.route("/pipeline")
@admin_required
def pipeline():
    api_key_set = bool(os.environ.get("OPENROUTER_API_KEY", "").strip())
    return render_template("pipeline.html", stats=get_stats(), api_key_set=api_key_set)


# ── Data API ───────────────────────────────────────────────────────────────────

@app.route("/api/stats")
@login_required
def api_stats():
    return jsonify(get_stats())


@app.route("/api/analytics")
@login_required
def api_analytics():
    return jsonify(get_analytics_data())


# ── Export API ─────────────────────────────────────────────────────────────────

@app.route("/api/geojson")
def api_geojson():
    """Public — used by the map iframe."""
    rows = get_all_observations()
    features = []
    for r in rows:
        try:
            lat = float(r.get("latitude_dd") or 0)
            lon = float(r.get("longitude_dd") or 0)
            if lat == 0 and lon == 0:
                continue
        except (ValueError, TypeError):
            continue
        features.append({
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": [lon, lat]},
            "properties": {
                "picture_name":    r.get("picture_name", ""),
                "category":        r["_category"],
                "species_name":    r.get("common_name") or r.get("species_name", ""),
                "scientific_name": r.get("scientific_name", ""),
                "date":            r.get("date", ""),
                "altitude":        r.get("altitude_m", ""),
                "cloudinary_url":  r.get("cloudinary_url", ""),
            },
        })
    geojson_str = json.dumps({"type": "FeatureCollection", "features": features},
                             ensure_ascii=False, indent=2)
    return Response(geojson_str, mimetype="application/geo+json",
                    headers={"Content-Disposition":
                             "attachment; filename=bio_observations.geojson"})


@app.route("/api/export-csv")
@login_required
def api_export_csv():
    cat_filter = request.args.get("category", "all")
    rows = get_all_observations(cat_filter)
    if not rows:
        return jsonify({"error": "No data"}), 404

    skip = {"_category", "id"}
    fieldnames = [k for k in rows[0].keys() if k not in skip]

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)

    fname = f"bio_{cat_filter}.csv"
    return Response(buf.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename={fname}"})


@app.route("/api/export-excel")
@login_required
def api_export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl not installed."}), 500

    cat_filter = request.args.get("category", "all")
    rows = get_all_observations(cat_filter)
    if not rows:
        return jsonify({"error": "No data"}), 404

    skip = {"_category", "id"}
    wb = openpyxl.Workbook()
    sheets_data = {}
    for r in rows:
        sheets_data.setdefault(r["_category"], []).append(r)

    CAT_COLORS = {"insect": "FFF59E0B", "flora": "FF4ADE80", "fungus": "FFC084FC"}

    for idx, (cat, cat_rows) in enumerate(sheets_data.items()):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = cat.capitalize()
        fieldnames = [k for k in cat_rows[0].keys() if k not in skip]
        hfill = PatternFill("solid", fgColor=CAT_COLORS.get(cat, "FF4ADE80"))
        hfont = Font(bold=True, color="FF000000")
        for col_i, field in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col_i, value=field.replace("_", " ").title())
            cell.fill = hfill; cell.font = hfont
            cell.alignment = Alignment(horizontal="center")
        for row_i, r in enumerate(cat_rows, 2):
            for col_i, field in enumerate(fieldnames, 1):
                ws.cell(row=row_i, column=col_i, value=r.get(field, ""))
        for col_i in range(1, len(fieldnames) + 1):
            ws.column_dimensions[get_column_letter(col_i)].auto_size = True

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"bio_{cat_filter}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/backup")
@login_required
def api_backup():
    rows = get_all_observations()
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for cat in CATEGORIES:
            cat_rows = [r for r in rows if r["_category"] == cat]
            if cat_rows:
                skip = {"_category", "id"}
                fieldnames = [k for k in cat_rows[0].keys() if k not in skip]
                sbuf = io.StringIO()
                writer = csv.DictWriter(sbuf, fieldnames=fieldnames, extrasaction="ignore")
                writer.writeheader(); writer.writerows(cat_rows)
                zf.writestr(f"bio_{cat}.csv", sbuf.getvalue())
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="biofield_backup.zip",
                     mimetype="application/zip")


# ── Image API ──────────────────────────────────────────────────────────────────

@app.route("/api/image-url/<int:obs_id>")
@login_required
def api_image_url(obs_id: int):
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT cloudinary_url FROM observations WHERE id = %s", (obs_id,))
            row = cur.fetchone()
    if not row or not row[0]:
        return "Not found", 404
    return jsonify({"url": row[0]})


# ── Upload API ─────────────────────────────────────────────────────────────────

@app.route("/api/upload", methods=["POST"])
@login_required
def api_upload():
    app.logger.info("Upload request received")
    category = request.form.get("category", "unsorted")
    files    = request.files.getlist("images")
    app.logger.info(f"Category: {category} | Files: {len(files)}")

    if category not in CATEGORIES:
        return jsonify({"status": "error", "message": f"Unknown category: {category}"}), 400
    if not files:
        return jsonify({"status": "error", "message": "No files received"}), 400

    saved_urls = []
    skipped    = []

    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                for f in files:
                    if not f or not f.filename:
                        continue
                    ext = "." + f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
                    if ext not in ALLOWED_EXT:
                        skipped.append(f.filename)
                        continue

                    result = cloudinary.uploader.upload(
                        f,
                        folder=f"biofield/{category}",
                        resource_type="image",
                    )
                    url       = result["secure_url"]
                    public_id = result["public_id"]
                    app.logger.info(f"Cloudinary OK: {url}")

                    cur.execute("""
                        INSERT INTO observations
                            (category, picture_name, cloudinary_url,
                             cloudinary_public_id, processing_status)
                        VALUES (%s, %s, %s, %s, 'PENDING')
                        RETURNING id
                    """, (category, f.filename, url, public_id))
                    obs_id = cur.fetchone()[0]

                    saved_urls.append({"url": url, "id": obs_id, "name": f.filename})

            conn.commit()

        return jsonify({
            "status":      "success",
            "total_saved": len(saved_urls),
            "urls":        saved_urls,
            "skipped":     skipped,
            "message":     "Images uploaded to Cloudinary and registered in database.",
        })

    except Exception as e:
        app.logger.error(f"Upload failed: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500


# ── Admin: Delete / Update observation ────────────────────────────────────────

@app.route("/api/observations/<int:obs_id>", methods=["DELETE"])
@admin_required
def api_delete_observation(obs_id: int):
    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT cloudinary_public_id FROM observations WHERE id = %s",
                    (obs_id,)
                )
                row = cur.fetchone()
                if not row:
                    return jsonify({"error": "Not found"}), 404
                public_id = row[0]
                cur.execute("DELETE FROM observations WHERE id = %s", (obs_id,))
            conn.commit()

        if public_id:
            try:
                cloudinary.uploader.destroy(public_id)
            except Exception as e:
                app.logger.warning(f"Cloudinary delete failed for {public_id}: {e}")

        return jsonify({"status": "deleted", "id": obs_id})
    except Exception as e:
        app.logger.error(f"Delete failed: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/observations/<int:obs_id>", methods=["PATCH"])
@admin_required
def api_update_observation(obs_id: int):
    data = request.get_json(silent=True) or {}
    allowed_fields = {
        "common_name", "scientific_name", "species_name", "category",
        "date", "gps_string", "latitude_dd", "longitude_dd",
        "altitude_m", "processing_status",
    }
    updates = {k: v for k, v in data.items() if k in allowed_fields}
    if not updates:
        return jsonify({"error": "No valid fields to update"}), 400

    set_clause = ", ".join(f"{k} = %s" for k in updates)
    values     = list(updates.values()) + [obs_id]

    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    f"UPDATE observations SET {set_clause} WHERE id = %s",
                    values
                )
            conn.commit()
        return jsonify({"status": "updated", "id": obs_id})
    except Exception as e:
        app.logger.error(f"Update failed: {e}")
        return jsonify({"error": str(e)}), 500


# ── Pipeline API ───────────────────────────────────────────────────────────────

@app.route("/api/run-extract", methods=["POST"])
@admin_required
def api_run_extract():
    import uuid
    sid = str(uuid.uuid4())
    _log_queues[sid] = queue.Queue()
    threading.Thread(target=_stream_subprocess,
                     args=([sys.executable, str(BASE_DIR / "extract.py")], sid),
                     daemon=True).start()
    return jsonify({"stream_id": sid})


@app.route("/api/run-map", methods=["POST"])
@admin_required
def api_run_map():
    import uuid
    sid = str(uuid.uuid4())
    _log_queues[sid] = queue.Queue()
    threading.Thread(target=_stream_subprocess,
                     args=([sys.executable, str(BASE_DIR / "map.py")], sid),
                     daemon=True).start()
    return jsonify({"stream_id": sid})


@app.route("/api/logs/<stream_id>")
@admin_required
def api_logs(stream_id: str):
    return Response(stream_with_context(sse_generator(stream_id)),
                    mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ── Startup ────────────────────────────────────────────────────────────────────

with app.app_context():
    try:
        init_db()
    except Exception as e:
        app.logger.error(f"init_db FAILED: {e}")
    try:
        seed_admin()
    except Exception as e:
        app.logger.error(f"seed_admin FAILED: {e}")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    app.run(host="0.0.0.0", port=port, threaded=True)
